from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
def result_in_excel(title_name,website):    
    # ws = wb.create_sheet(title='1st-Sem')
    ws = wb.create_sheet(title=title_name) # title_name
    ws.sheet_properties.tabColor = "AB0000" # RGB in hexadecimal format
    ws['A1'] = 'Roll-No'
    ws['B1'] = 'GPA'

    for i in [9,19,24,49,59]:
        driver = webdriver.Chrome('C:\\Users\\user\\Downloads\\chromedriver')
        # driver.get('https://www.osmania.ac.in/res07/20200270.jsp')
        driver.get(website) # website
        wait= WebDriverWait(driver, 60)
        stdstr = '245318733'
        input_box= driver.find_element_by_name('htno')
        if(i<10):
            res_string = stdstr + '00' + str(i)
            input_box.send_keys(res_string+Keys.ENTER)
            ws['A'+str(eval('i+1'))] = res_string
            wait= WebDriverWait(driver, 20)
            # tablereq = driver.find_element_by_xpath("//*[@id='AutoNumber5']/tbody/tr["+str(3)+"]/td["+str(2)+"]/b/font").text
            tablereq = driver.find_element_by_xpath("//*[@id='AutoNumber5']/tbody/tr["+str(3)+"]/td["+str(3)+"]/b/font").text
            print(tablereq)
            ws['B'+str(eval('i+1'))]=tablereq
        else:
            res_string = stdstr + '0' + str(i)
            input_box.send_keys(res_string+Keys.ENTER)
            ws['A'+str(eval('i+1'))] = res_string
            # tablereq = driver.find_element_by_xpath("//*[@id='AutoNumber5']/tbody/tr["+str(3)+"]/td["+str(2)+"]/b/font").text
            tablereq = driver.find_element_by_xpath("//*[@id='AutoNumber5']/tbody/tr["+str(3)+"]/td["+str(3)+"]/b/font").text
            print(tablereq)
            ws['B'+str(eval('i+1'))]=tablereq
        driver.close()    
# wb.save('sem1_res.xlsx')
"""
//*[@id="AutoNumber5"]/tbody/tr["+str(3)+"]/td["+str(2)+"]/b/font
"""
"""
//*[@id="AutoNumber5"]/tbody/tr["+str(3)+"]/td["+str(3)+"]/b/font
"""
result_in_excel('1sem','https://www.osmania.ac.in/res07/20190318.jsp')
result_in_excel('2sem','https://www.osmania.ac.in/res07/20190855.jsp')
# result_in_excel('3sem','https://www.osmania.ac.in/res07/20200270.jsp')
wb.save('sem12_res.xlsx') # saved_sheetname 
